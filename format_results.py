import itertools,os,json
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment,Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import string
import re
from sqlalchemy.orm.collections import collection
from surrealdb import Surreal

from surreal import create_conn

db = Surreal('https://aws.datahubweb.com:6000')
db.signin({'username': "root", 'password': os.getenv("SURREALDB_PASSWORD")})
db.use("llm_study", "benchmark")
rows=db.query("select * from prompt")
data=pd.DataFrame.from_records(rows)
# data.to_csv('/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/data/prompt_responses.csv',index=False)
original_data=pd.read_excel('/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/data/Consolidated Dataset V7.xlsx')

# format results of deep seek to remove the <think/> tags
def format_deepseek(text):
    formated_string=re.sub(r'<think>.*?</think>\n*', '', text, flags=re.DOTALL)
    return formated_string

data['response']=data['response'].map(format_deepseek)

prompt_sys='prompt1'
data=data.loc[data['system_prompt']==prompt_sys].copy()
# data=pd.read_csv('/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/results/prompt_output.csv')
# data['system_prompt']=data['system_prompt'].fillna('')
letters=string.ascii_uppercase
# model_names=data['model'].unique()
# model_names=['gemma3:4b','llama3.2:3b', 'deepseek-r1:8b']
model_names=['medgemma', 'gemini-2.5-flash', 'deepSeek-r1', 'gpt-4.1', 'o3', ]


prompt_strings=data['user_prompt'].unique()
system_prompt_strings=data['system_prompt'].unique()

## assign and randomize
prompt_list=data['index'].unique().tolist()
sorted(prompt_list)
prompt_list=sorted(prompt_list)
model_order={}
rng=np.random.RandomState(42)
for ind in prompt_list: model_order[ind] = rng.choice(model_names,size=len(model_names),replace=False).tolist()

with open('/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/data/shuffled_model_order.json', 'w') as outfile:
    json.dump(model_order, outfile,indent=4)


n_batches=list(itertools.batched(prompt_list,15))
with open('/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/data/batches_master_index.json', 'w') as outfile:
    json.dump({f'batch{i+1}':b for i,b in enumerate(n_batches)}, outfile,indent=4)

##format results
for b,b_index in enumerate(n_batches):
    wb=openpyxl.Workbook()
    ws = wb.active
    ws.append(['StudyID','User Prompt','Clinician response'] + [f'Model {m+1}' for m in range(len(model_names))])

    for row, b_index_ in enumerate(b_index):
        sub=data.loc[data['index']==b_index_]
        ws[f'A{row + 2}'] = b_index_
        ws[f'B{row + 2}'] =sub['user_prompt'].iat[0]
        ws[f'C{row + 2}'] = original_data.loc[original_data['Master_Index']==b_index_,'Clinician'].iat[0]
        for m_string_i, m_string in enumerate(model_order[b_index_]):
            cell_data=sub.loc[sub['model']==m_string,'response'].iloc[0]
            ws[f'{letters[m_string_i+3]}{row + 2}'] = cell_data

    #change column widths
    for c in range(1,len(model_names)+3):
        ws.column_dimensions[string.ascii_uppercase[c]].width = 70
    #
    c = ws['C2']
    ws.freeze_panes = c
    #
    # # for column in ws.iter_cols():
    # #     column.alignment = Alignment(horizontal='center', vertical='center')
    #
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(f"/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/data/batches/batch{b+1}.xlsx")




wb=openpyxl.Workbook()
for m in model_names:
    prompt_data=data.loc[data['model']==m]

    ws= wb.create_sheet(title=m.replace(":","-"),index=0)
    ws.append(['User Prompt',]+system_prompt_strings.tolist())
    for row,user_p in enumerate(prompt_strings):
        ws[f'A{row+2}']=user_p
        for col,s_prompt in enumerate(system_prompt_strings,1):
            cell_data=prompt_data.loc[(prompt_data['system_prompt']==s_prompt) & (prompt_data['user_prompt']==user_p),'response'].iat[0]
            ws[f'{letters[col]}{row+2}']=cell_data

    #change column widths
    for c in range(len(system_prompt_strings)+1):
        ws.column_dimensions[string.ascii_uppercase[c]].width = 70

    c = ws['B2']
    ws.freeze_panes = c

    # for column in ws.iter_cols():
    #     column.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(f"/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/results/formated prompt outputs.xlsx")


# All prompt responses in one worksheet
##format results
wb=openpyxl.Workbook()
ws = wb.active
ws.append(['StudyID','User Prompt','Clinician response'] + model_names)

for row, b_index_ in enumerate(data['index'].unique()):
    sub=data.loc[data['index']==b_index_]
    ws[f'A{row + 2}'] = b_index_
    ws[f'B{row + 2}'] =sub['user_prompt'].iat[0]
    ws[f'C{row + 2}'] = original_data.loc[original_data['Master_Index']==b_index_,'Clinician'].iat[0]
    for m_string_i, m_string in enumerate(model_names):
        cell_data=sub.loc[sub['model']==m_string,'response'].iloc[0]
        ws[f'{letters[m_string_i+3]}{row + 2}'] = cell_data

#change column widths
for c in range(1,len(model_names)+3):
    ws.column_dimensions[string.ascii_uppercase[c]].width = 70
#
c = ws['C2']
ws.freeze_panes = c
#
# # for column in ws.iter_cols():
# #     column.alignment = Alignment(horizontal='center', vertical='center')
#
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(f"/home/pmwaniki/Dropbox/others/Ambrose/Sync study/Benchmark/data/Prompt responses.xlsx")
